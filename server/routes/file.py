from flask import Blueprint, request, jsonify, send_from_directory, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from werkzeug.utils import secure_filename
from models import db, File, User
import os
import uuid
import hashlib
import magic
import PyPDF2
import docx
import openpyxl
import openai
import logging
import json

file_bp = Blueprint('file_bp', __name__)

# Helper functions (can be moved to a separate utils file if more are added)
def get_cached_response(key):
    redis_client = current_app.config['SESSION_REDIS']
    cached_data = redis_client.get(key)
    if cached_data:
        return json.loads(cached_data)
    return None

def set_cached_response(key, data, ex=3600): # Cache for 1 hour by default
    redis_client = current_app.config['SESSION_REDIS']
    redis_client.setex(key, ex, json.dumps(data))


@file_bp.route('/upload-file', methods=['POST'])
@jwt_required()
def upload_file():
    if 'file' not in request.files or 'name' not in request.form:
        logging.warning("Upload file attempt with missing file or name.")
        return jsonify({"error": "No file or name provided"}), 400
    
    file = request.files['file']
    name = request.form['name']
    
    if not file or not name:
        logging.warning("Upload file attempt with invalid file or name.")
        return jsonify({"error": "Invalid file or name"}), 400

    # Secure filename and generate unique name
    original_filename = secure_filename(file.filename)
    file_extension = os.path.splitext(original_filename)[1]
    unique_filename = f"{uuid.uuid4().hex}{file_extension}"
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)
    file.save(filepath)

    # Content type validation using python-magic
    mime_type = magic.from_file(filepath, mime=True)
    allowed_mimetypes = [
        'application/pdf',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ]
    if mime_type not in allowed_mimetypes:
        os.remove(filepath) # Delete the uploaded file if it's not allowed
        logging.warning(f"Unsupported file MIME type uploaded: {mime_type}")
        return jsonify({"error": f"Unsupported file MIME type: {mime_type}"}), 400
    
    user_id = get_jwt_identity()
    if not user_id:
        logging.error("Unauthorized access to upload-file route without valid JWT.")
        return jsonify({"error": "Unauthorised"}), 401
    
    # Extract text from file
    text_content = ""
    try:
        if file_extension == '.pdf':
            reader = PyPDF2.PdfReader(filepath)
            for page in reader.pages:
                text_content += page.extract_text()
        elif file_extension == '.docx':
            doc = docx.Document(filepath)
            for para in doc.paragraphs:
                text_content += para.text + "\n"
        elif file_extension == '.xlsx':
            workbook = openpyxl.load_workbook(filepath)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            text_content += str(cell.value) + "\t"
                    text_content += "\n"
        else:
            logging.error(f"File extension {file_extension} not handled for text extraction.")
            return jsonify({"error": "Unsupported file type for text extraction"}), 400
    except Exception as e:
        logging.exception(f"Error processing file {filepath} for text extraction.")
        return jsonify({"error": f"Error processing file: {str(e)}"}), 500

    # Generate cache key for OpenAI calls
    cache_key_prefix = hashlib.md5(text_content.encode('utf-8')).hexdigest()
    summary_cache_key = f"summary:{cache_key_prefix}"
    questions_cache_key = f"questions:{cache_key_prefix}"

    summary = get_cached_response(summary_cache_key)
    questions = get_cached_response(questions_cache_key)

    if not summary:
        try:
            # OpenAI call for summarization
            summary_response = openai.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are a helpful assistant."},
                    {"role": "user", "content": f"Summarize the following text: {text_content}"}
                ]
            )
            summary = summary_response.choices[0].message.content
            set_cached_response(summary_cache_key, summary)
        except Exception as e:
            logging.exception(f"Error generating summary for file {filepath}.")
            return jsonify({"error": f"Error generating summary: {str(e)}"}), 500

    if not questions:
        try:
            # OpenAI call for question generation
            questions_response = openai.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are a helpful assistant."},
                    {"role": "user", "content": f"Generate 5 multiple-choice questions based on the following text: {text_content}"}
                ]
            )
            questions = questions_response.choices[0].message.content
            set_cached_response(questions_cache_key, questions)
        except Exception as e:
            logging.exception(f"Error generating questions for file {filepath}.")
            return jsonify({"error": f"Error generating questions: {str(e)}"}), 500
    
    new_file = File(name=name, path=filepath, uploaded_by=user_id, summary=summary, questions=questions)
    db.session.add(new_file)
    db.session.commit()
    
    response = jsonify({"message": "File uploaded successfully", "file": {"name": name, "path": filepath, "summary": summary, "questions": questions}})
    response.headers.add("Access-Control-Allow-Origin", "http://localhost:5173")
    response.headers.add("Access-Control-Allow-Credentials", "true")
    return response, 201


@file_bp.route('/search-files', methods=['GET'])
@jwt_required()
def search_files():
    user_id = get_jwt_identity()
    if not user_id:
        logging.error("Unauthorized access to search-files route without valid JWT.")
        return jsonify({"error": "Unauthorised"}), 401

    query = request.args.get('query', '')
    if not query:
        logging.warning("Search files attempt with no query provided.")
        return jsonify({"error": "No search query provided"}), 400

    files = File.query.filter(File.name.ilike(f"%{query}%", File.uploaded_by == user_id)).all()
    results = [{"id": file.id, "name": file.name, "path": file.path, "summary": file.summary, "questions": file.questions} for file in files]

    return jsonify({"files": results}), 200

@file_bp.route('/files/<filename>', methods=['GET'])
@jwt_required()
def serve_file(filename):
    user_id = get_jwt_identity()
    if not user_id:
        logging.error("Unauthorized access to serve-file route without valid JWT.")
        return jsonify({"error": "Unauthorised"}), 401

    # Ensure the filename is secure to prevent directory traversal
    safe_filename = secure_filename(filename)

    # Retrieve file information from the database
    redis_client = current_app.config['SESSION_REDIS']
    file_record = File.query.filter_by(path=os.path.join(current_app.config['UPLOAD_FOLDER'], safe_filename), uploaded_by=user_id).first()

    if not file_record:
        logging.warning(f"File not found or unauthorized: {safe_filename} for user {user_id}")
        return jsonify({"error": "File not found or not authorized"}), 404

    try:
        return send_from_directory(current_app.config['UPLOAD_FOLDER'], safe_filename)
    except Exception as e:
        logging.exception(f"Error serving file {safe_filename}.")
        return jsonify({"error": f"Error serving file: {str(e)}"}), 500